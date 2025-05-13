import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Jira Tasks"

# Define headers based on the table structure in HTML
headers = [
    "Summary",
    "Description",
    "Epic",
    "Priority",
    "Estimate Value",
    "Estimate Unit",
    "Story Points",
    "Start Date (YYYY-MM-DD)",
    "Due Date (YYYY-MM-DD)",
    "Notes"
]

# Apply header styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

    # Set column width based on the header content
    ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(header) + 5)

# Add input validation for different fields
# Priority dropdown
priority_validation = DataValidation(
    type="list",
    formula1='"Highest,High,Medium,Low,Lowest"',
    allow_blank=True
)
priority_validation.error = "Please select a valid priority"
priority_validation.errorTitle = "Invalid Priority"
priority_validation.prompt = "Select priority from dropdown"
priority_validation.promptTitle = "Priority"
ws.add_data_validation(priority_validation)
priority_validation.add('D2:D1000')  # Apply to column D from row 2 to 1000

# Estimate Unit dropdown
estimate_unit_validation = DataValidation(
    type="list",
    formula1='"m,h,d"',
    allow_blank=True
)
estimate_unit_validation.error = "Please select a valid unit (m=minutes, h=hours, d=days)"
estimate_unit_validation.errorTitle = "Invalid Unit"
estimate_unit_validation.prompt = "m=minutes, h=hours, d=days"
estimate_unit_validation.promptTitle = "Unit"
ws.add_data_validation(estimate_unit_validation)
estimate_unit_validation.add('F2:F1000')  # Apply to column F from row 2 to 1000

# Story Points validation (numeric value)
story_points_validation = DataValidation(
    type="decimal",
    operator="greaterThanOrEqual",
    formula1="0",
    allow_blank=True
)
story_points_validation.error = "Story points must be a non-negative number"
story_points_validation.errorTitle = "Invalid Story Points"
ws.add_data_validation(story_points_validation)
story_points_validation.add('G2:G1000')  # Apply to column G from row 2 to 1000

# Estimate Value validation (numeric value)
estimate_value_validation = DataValidation(
    type="decimal",
    operator="greaterThanOrEqual",
    formula1="0",
    allow_blank=True
)
estimate_value_validation.error = "Estimate value must be a non-negative number"
estimate_value_validation.errorTitle = "Invalid Estimate"
ws.add_data_validation(estimate_value_validation)
estimate_value_validation.add('E2:E1000')  # Apply to column E from row 2 to 1000

# Date validation format - not strict but provides guidance
# Note: We use info/prompt to guide users on the date format
for col_letter in ['H', 'I']:  # Start and Due Date columns
    for row_idx in range(2, 1001):
        cell = ws.cell(row=row_idx, column=ord(col_letter) - ord('A') + 1)
        cell.number_format = 'yyyy-mm-dd'

# Add instructions in a separate sheet
instructions_sheet = wb.create_sheet(title="Instructions")
instructions = [
    ["Jira Task Import Template - Instructions"],
    [""],
    ["Field", "Description", "Format/Options"],
    ["Summary", "Task title/name (Required)", "Text"],
    ["Description", "Detailed description of the task", "Text"],
    ["Epic", "Name of Epic to link the task to", "Text"],
    ["Priority", "Task priority level", "Highest, High, Medium, Low, Lowest"],
    ["Estimate Value", "Numeric value of time estimate", "Positive number"],
    ["Estimate Unit", "Unit for time estimate", "m (minutes), h (hours), d (days)"],
    ["Story Points", "Story points for the task", "Positive number"],
    ["Start Date", "Task start date", "YYYY-MM-DD"],
    ["Due Date", "Task due date", "YYYY-MM-DD"],
    ["Notes", "Additional notes", "Text"],
    [""],
    ["Note: After import, you'll be able to edit tasks before creating them in Jira."]
]

# Add instructions with formatting
instruction_title_font = Font(bold=True, size=14)
header_font_instructions = Font(bold=True)

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = instructions_sheet.cell(row=row_idx, column=col_idx)
        cell.value = value

        # Apply formatting
        if row_idx == 1:  # Title
            cell.font = instruction_title_font
        elif row_idx == 3:  # Header row
            cell.font = header_font_instructions
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Set column widths in instructions sheet
instructions_sheet.column_dimensions['A'].width = 15
instructions_sheet.column_dimensions['B'].width = 35
instructions_sheet.column_dimensions['C'].width = 25

# Add example data row
example_data = [
    "Implement login feature",
    "Create login page with username/password fields",
    "User Authentication",
    "Medium",
    "8",
    "h",
    "5",
    "2023-05-01",
    "2023-05-10",
    "Follow design mockup in Figma"
]

for col_idx, value in enumerate(example_data, 1):
    ws.cell(row=2, column=col_idx).value = value

# Create directory if it doesn't exist
os.makedirs('templates/static', exist_ok=True)

# Save the template
wb.save('templates/static/jira_tasks_template.xlsx')
print("Excel template created successfully at templates/static/jira_tasks_template.xlsx")
